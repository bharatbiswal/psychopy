#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import absolute_import, print_function
import sys

from pkg_resources import parse_version

from .base import DataHandler
from .experiment import ExperimentHandler
from .trial import TrialHandler, TrialHandler2, TrialHandlerExt, TrialType
from .staircase import (StairHandler, QuestHandler, PsiHandler,
                        MultiStairHandler)

if sys.version_info.major == 3 and sys.version_info.minor >= 6:
    from .staircase import QuestPlusWeibullHandler

from .utils import (checkValidFilePath, isValidVariableName, importTrialTypes,
                    sliceFromString, indicesFromString, importConditions,
                    createFactorialTrialList, bootStraps, functionFromStaircase,
                    getDateStr)

from .fit import (FitFunction, FitCumNormal, FitLogistic, FitNakaRushton,
                  FitWeibull)

try:
    # import openpyxl
    import openpyxl
    if parse_version(openpyxl.__version__) >= parse_version('2.4.0'):
        # openpyxl moved get_column_letter to utils.cell
        from openpyxl.utils.cell import get_column_letter
    else:
        from openpyxl.cell import get_column_letter
    from openpyxl.reader.excel import load_workbook
    haveOpenpyxl = True
except ImportError:
    haveOpenpyxl = False

try:
    import xlrd
    haveXlrd = True
except ImportError:
    haveXlrd = False
